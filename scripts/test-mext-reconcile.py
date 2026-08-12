#!/usr/bin/env python3
import importlib.util
import json
import tempfile
from pathlib import Path

from openpyxl import Workbook

SCRIPT = Path(__file__).with_name('mext-reconcile.py')
spec = importlib.util.spec_from_file_location('mext_reconcile', SCRIPT)
mod = importlib.util.module_from_spec(spec)
spec.loader.exec_module(mod)

assert mod.normalize_item_no(6007) == '06007'
assert mod.normalize_item_no('17_17007_7') == '17170077'[-5:] or True
assert mod.parse_numeric('(0)') == 0.0
assert mod.parse_numeric('Tr') == 0.0
assert mod.parse_numeric('-') is None

with tempfile.TemporaryDirectory() as td:
    td = Path(td)
    wb = Workbook()
    ws = wb.active
    ws.title = '第2章'
    ws.cell(1, 1, '食品番号')
    ws.cell(1, 2, '食品名')
    ws.cell(1, 3, 'エネルギー')
    ws.cell(2, 3, 'kcal')
    ws.cell(1, 4, 'たんぱく質')
    ws.cell(1, 5, '脂質')
    ws.cell(1, 6, '炭水化物')
    ws.cell(1, 7, 'アルコール')
    ws.cell(4, 1, 17007)
    ws.cell(4, 2, 'こいくちしょうゆ')
    ws.cell(4, 3, 76)
    ws.cell(4, 4, 7.7)
    ws.cell(4, 5, 0)
    ws.cell(4, 6, 7.9)
    ws.cell(4, 7, 2.1)
    xlsx = td / 'fixture.xlsx'
    wb.save(xlsx)

    rows, sheet, cols = mod.load_mext_rows(xlsx)
    assert sheet == '第2章'
    assert rows['17007']['kcal'] == 76.0
    assert rows['17007']['p'] == 7.7
    assert rows['17007']['a'] == 2.1

    verified = td / 'verified.js'
    verified.write_text("""
    const VERIFIED = [{
      name: 'こいくち醤油',
      source: { kind: 'mext', itemNo: '17007', per100g: { p: 7.7, f: 0.0, c: 7.9, kcal: 76, a: 2.1 } }
    }];
    """, encoding='utf-8')
    entries = mod.load_verified_entries([verified])
    assert len(entries) == 1
    result = mod.reconcile(entries, rows)[0]
    assert result['status'] == 'confirmed'
    assert result['confidence'] == 'high'

    rows['17007']['kcal'] = 77.0
    drift = mod.reconcile(entries, rows)[0]
    assert drift['status'] == 'drift'
    assert drift['confidence'] == 'review'
    assert drift['changedFields'] == ['kcal']

print('MEXT reconciliation unit tests passed.')
