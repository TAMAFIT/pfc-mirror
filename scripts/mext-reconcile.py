#!/usr/bin/env python3
import argparse
import hashlib
import html.parser
import json
import math
import re
import sys
import urllib.parse
import urllib.request
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

MEXT_PAGE = "https://www.mext.go.jp/a_menu/syokuhinseibun/mext_00001.html"
USER_AGENT = "TAMAFIT-PFC-FoodMaster-Reconciler/1.0 (+https://github.com/TAMAFIT/pfc-mirror)"


class LinkParser(html.parser.HTMLParser):
    def __init__(self):
        super().__init__()
        self.links = []
        self._href = None
        self._text = []

    def handle_starttag(self, tag, attrs):
        if tag.lower() == "a":
            self._href = dict(attrs).get("href")
            self._text = []

    def handle_data(self, data):
        if self._href is not None:
            self._text.append(data)

    def handle_endtag(self, tag):
        if tag.lower() == "a" and self._href is not None:
            self.links.append((self._href, "".join(self._text).strip()))
            self._href = None
            self._text = []


def request_bytes(url):
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(req, timeout=45) as response:
        return response.read(), response.headers.get("Content-Type", "")


def discover_workbook_url(page_url=MEXT_PAGE):
    body, _ = request_bytes(page_url)
    parser = LinkParser()
    parser.feed(body.decode("utf-8", errors="replace"))
    candidates = []
    for href, text in parser.links:
        if not href:
            continue
        absolute = urllib.parse.urljoin(page_url, href)
        if not absolute.lower().endswith((".xlsx", ".xls")):
            continue
        normalized = re.sub(r"\s+", "", text)
        score = 0
        if "第2章" in normalized:
            score += 20
        if "データ" in normalized:
            score += 10
        if all(x not in normalized for x in ("第1表", "第2表", "第3表", "第4表", "別表")):
            score += 5
        if "正誤" in normalized:
            score -= 100
        candidates.append((score, absolute, text))
    if not candidates:
        raise RuntimeError("MEXT main-table Excel link could not be discovered")
    candidates.sort(reverse=True)
    score, url, text = candidates[0]
    if score < 20:
        raise RuntimeError(f"MEXT workbook discovery confidence too low: {text} -> {url}")
    return url, text


def normalize_item_no(value):
    if value is None:
        return ""
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if isinstance(value, float) and not value.is_integer():
            return ""
        return str(int(value)).zfill(5)
    raw = re.sub(r"\D", "", str(value))
    return raw.zfill(5) if raw else ""


def parse_numeric(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    raw = str(value).strip().replace("＊", "").replace("*", "")
    if raw in {"-", "…", "..."}:
        return None
    if raw in {"Tr", "tr", "(Tr)", "(tr)"}:
        return 0.0
    if raw.startswith("(") and raw.endswith(")"):
        raw = raw[1:-1].strip()
    raw = raw.replace(",", "")
    try:
        return float(raw)
    except ValueError:
        return None


def cell_text(value):
    return "" if value is None else str(value).strip()


def pick_sheet_and_columns(workbook):
    best = None
    for ws in workbook.worksheets:
        max_probe_rows = min(ws.max_row, 35)
        max_probe_cols = min(ws.max_column, 140)
        col_tokens = {}
        col_values = {}
        for c in range(1, max_probe_cols + 1):
            vals = [cell_text(ws.cell(r, c).value) for r in range(1, max_probe_rows + 1)]
            vals = [v for v in vals if v]
            col_values[c] = vals
            col_tokens[c] = " ".join(vals)

        def score_col(predicate):
            hits = []
            for c, token in col_tokens.items():
                score = predicate(token, col_values[c])
                if score:
                    hits.append((score, c, token))
            hits.sort(reverse=True)
            return hits[0] if hits else None

        food_no = score_col(lambda t, v: 100 if "食品番号" in t else 0)
        food_name = score_col(lambda t, v: 100 if "食品名" in t else 0)
        kcal = score_col(lambda t, v: 130 if "エネルギー" in t and "kcal" in t.lower() else 0)
        protein = score_col(lambda t, v: 130 if any(x == "たんぱく質" for x in v) and "アミノ酸組成" not in t else 0)
        fat = score_col(lambda t, v: 130 if any(x == "脂質" for x in v) and "脂肪酸" not in t else 0)
        carbs = score_col(lambda t, v: 130 if any(x == "炭水化物" for x in v) and "利用可能炭水化物" not in t else 0)
        alcohol = score_col(lambda t, v: 110 if any(x == "アルコール" for x in v) else 0)
        required = [food_no, food_name, kcal, protein, fat, carbs]
        if all(required):
            total = sum(x[0] for x in required) + (alcohol[0] if alcohol else 0)
            candidate = (total, ws, {
                "food_no": food_no[1], "food_name": food_name[1], "kcal": kcal[1],
                "p": protein[1], "f": fat[1], "c": carbs[1],
                "a": alcohol[1] if alcohol else None,
            })
            if best is None or candidate[0] > best[0]:
                best = candidate
    if best is None:
        raise RuntimeError("Could not locate required MEXT nutrient columns in workbook")
    return best[1], best[2]


def load_mext_rows(xlsx_path):
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws, cols = pick_sheet_and_columns(wb)
    rows = {}
    for r in range(1, ws.max_row + 1):
        item_no = normalize_item_no(ws.cell(r, cols["food_no"]).value)
        if len(item_no) != 5 or item_no == "00000":
            continue
        name = cell_text(ws.cell(r, cols["food_name"]).value)
        values = {
            "p": parse_numeric(ws.cell(r, cols["p"]).value),
            "f": parse_numeric(ws.cell(r, cols["f"]).value),
            "c": parse_numeric(ws.cell(r, cols["c"]).value),
            "kcal": parse_numeric(ws.cell(r, cols["kcal"]).value),
            "a": parse_numeric(ws.cell(r, cols["a"]).value) if cols["a"] else 0.0,
        }
        if all(values[k] is None for k in ("p", "f", "c", "kcal")):
            continue
        rows[item_no] = {"itemNo": item_no, "officialName": name, **values}
    return rows, ws.title, cols


ENTRY_RE = re.compile(
    r"name:\s*'(?P<name>[^']+)'[\s\S]*?source:\s*\{\s*kind:\s*'mext'[\s\S]*?itemNo:\s*'(?P<item>\d+)'[\s\S]*?per100g:\s*\{\s*p:\s*(?P<p>-?\d+(?:\.\d+)?),\s*f:\s*(?P<f>-?\d+(?:\.\d+)?),\s*c:\s*(?P<c>-?\d+(?:\.\d+)?),\s*kcal:\s*(?P<kcal>-?\d+(?:\.\d+)?),\s*a:\s*(?P<a>-?\d+(?:\.\d+)?)\s*\}",
    re.MULTILINE,
)


def load_verified_entries(paths):
    entries = []
    for path in paths:
        text = Path(path).read_text(encoding="utf-8")
        for m in ENTRY_RE.finditer(text):
            entries.append({
                "file": str(path), "name": m.group("name"), "itemNo": normalize_item_no(m.group("item")),
                "per100g": {k: float(m.group(k)) for k in ("p", "f", "c", "kcal", "a")},
            })
    if not entries:
        raise RuntimeError("No MEXT verified entries were parsed from source files")
    return entries


def compare_value(key, stored, official):
    if official is None:
        return {"match": False, "reason": "official-missing", "stored": stored, "official": None}
    tolerance = 0.51 if key == "kcal" else 0.051
    match = abs(float(stored) - float(official)) <= tolerance
    return {"match": match, "stored": stored, "official": official, "delta": round(float(official) - float(stored), 4)}


def reconcile(entries, official_rows):
    results = []
    for entry in entries:
        official = official_rows.get(entry["itemNo"])
        if not official:
            results.append({**entry, "status": "missing-item", "confidence": "review", "official": None, "differences": {}})
            continue
        diffs = {k: compare_value(k, entry["per100g"][k], official.get(k)) for k in ("p", "f", "c", "kcal", "a")}
        bad = [k for k, v in diffs.items() if not v["match"]]
        results.append({
            **entry,
            "status": "confirmed" if not bad else "drift",
            "confidence": "high" if not bad else "review",
            "official": official,
            "differences": diffs,
            "changedFields": bad,
        })
    return results


def make_markdown(report):
    lines = [
        "# MEXT Food Master reconciliation",
        "",
        f"- Checked: `{report['checked']}`",
        f"- Confirmed: `{report['confirmed']}`",
        f"- Drift: `{report['drift']}`",
        f"- Missing item numbers: `{report['missing']}`",
        f"- Official workbook: `{report['source']['workbookUrl']}`",
        f"- Workbook SHA-256: `{report['source']['sha256']}`",
        "",
    ]
    changed = [x for x in report["entries"] if x["status"] != "confirmed"]
    if not changed:
        lines.append("All MEXT-backed Food Master entries match the current official workbook.")
        return "\n".join(lines) + "\n"
    lines += ["| Food | Item No | Status | Changed |", "|---|---:|---|---|"]
    for x in changed:
        fields = ", ".join(x.get("changedFields", [])) or "-"
        lines.append(f"| {x['name']} | {x['itemNo']} | {x['status']} | {fields} |")
    return "\n".join(lines) + "\n"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--verified", action="append", required=True)
    parser.add_argument("--xlsx")
    parser.add_argument("--out-json", default="mext-reconcile-report.json")
    parser.add_argument("--out-md", default="mext-reconcile-report.md")
    parser.add_argument("--strict", action="store_true")
    args = parser.parse_args()

    workbook_url = args.xlsx
    link_text = "explicit"
    temp_path = None
    if workbook_url and Path(workbook_url).exists():
        xlsx_path = Path(workbook_url)
        workbook_url = f"file://{xlsx_path.resolve()}"
    else:
        if not workbook_url:
            workbook_url, link_text = discover_workbook_url()
        body, content_type = request_bytes(workbook_url)
        if not body.startswith(b"PK"):
            raise RuntimeError(f"Downloaded MEXT workbook is not XLSX: {content_type}")
        temp_path = Path(".mext-food-master.xlsx")
        temp_path.write_bytes(body)
        xlsx_path = temp_path

    sha256 = hashlib.sha256(xlsx_path.read_bytes()).hexdigest()
    official_rows, sheet_title, columns = load_mext_rows(xlsx_path)
    entries = load_verified_entries(args.verified)
    results = reconcile(entries, official_rows)
    confirmed = sum(x["status"] == "confirmed" for x in results)
    drift = sum(x["status"] == "drift" for x in results)
    missing = sum(x["status"] == "missing-item" for x in results)
    report = {
        "schemaVersion": 1,
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "checked": len(results), "confirmed": confirmed, "drift": drift, "missing": missing,
        "source": {
            "pageUrl": MEXT_PAGE,
            "workbookUrl": workbook_url,
            "linkText": link_text,
            "sha256": sha256,
            "sheet": sheet_title,
            "columns": columns,
            "officialFoodRows": len(official_rows),
            "basis": "edible portion per 100g",
        },
        "entries": results,
    }
    Path(args.out_json).write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    Path(args.out_md).write_text(make_markdown(report), encoding="utf-8")

    print(f"MEXT_RECONCILE checked={len(results)} confirmed={confirmed} drift={drift} missing={missing}")
    print(f"MEXT_WORKBOOK {workbook_url}")
    print(f"MEXT_SHA256 {sha256}")
    for x in results:
        if x["status"] != "confirmed":
            print(f"MEXT_DIFF {x['name']} item={x['itemNo']} status={x['status']} fields={','.join(x.get('changedFields', []))}")

    if temp_path and temp_path.exists():
        temp_path.unlink()
    if args.strict and (drift or missing):
        return 2
    return 0


if __name__ == "__main__":
    sys.exit(main())
